import pprint
import os
import openpyxl
import re
from openpyxl.comments import Comment
from win32com.client import Dispatch
#python -m pip install pywin32
import hours_config


print('docha (d), tati (нажать любую клавишу):')
who = input()
if who == 'd':
    your_target_folder = hours_config.your_target_folder_docha
else:
    your_target_folder = hours_config.your_target_folder



def just_open(filename):
    xlApp = Dispatch('Excel.Application')
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()


def define_of_files(your_target_folder):
    for dirpath, _, filenames in os.walk(your_target_folder):
        fs_file_ = re.compile(r'.*(FS).*')
        stma_file_ = re.compile(r'.*(STEMA).*')
        rmtp_file_ = re.compile(r'^(rmtp).*')
        for items in filenames:
            if items.lower().endswith('.xlsx'):
                fs_file__ = fs_file_.match(items)
                if fs_file__:
                    fs_file = fr'{your_target_folder}/{fs_file__[0]}'

                stma_file__ = stma_file_.match(items)
                if stma_file__:
                    stma_file = fr'{your_target_folder}/{stma_file__[0]}'

                rmtp_file__ = rmtp_file_.match(items)
                if rmtp_file__:
                    rmtp_file = fr'{your_target_folder}/{rmtp_file__[0]}'
    return fs_file, stma_file, rmtp_file


def read_hours(file):
    ttl_hour_book = openpyxl.load_workbook(filename=file, data_only=True)
    ttl_hours = dict()
    sh_name = ttl_hour_book.sheetnames[0]
    sh = ttl_hour_book[sh_name]
    head_row = 5
    first_col_week = 43
    last_col_week = 47
    print(os.path.basename(file))
    pprint.pprint(list(cell.value for cell in sh[head_row][first_col_week - 1: last_col_week + 2]))
    for rx in range(head_row + 1, sh.max_row):
        fs_name = sh[rx][0]
        if fs_name.value:
            if len(fs_name.value.split()) >= 2:
                range_hours = sh[rx][first_col_week: last_col_week + 1]
                if fs_name.value:
                    list_hours = []
                    for l_ in range_hours:
                        list_hours.append(l_.value)
                    ttl_hours[fs_name.value.strip()] = list_hours
    return ttl_hours


def processing_comments(cell, comm, fs_time_i, rmtp_name):
    if cell.value:
        if cell.value != fs_time_i and not (cell.value is None or cell.value == 0):
            if comm:  # если у ячейки уже был комментарий
                comm = Comment(f'{comm.text}, было: {cell.value}, стало: {fs_time_i}', 'TJ')
            else:  # если комментов не было, но предыдущее значение ячейки не 0 и не None
                comm = Comment(f'было: {cell.value}, стало: {fs_time_i}', 'TJ')
                print(rmtp_name.value, cell.coordinate, comm.text)
                cell.comment = comm
                cell.fill = openpyxl.styles.PatternFill(start_color='FFC7CE', fill_type="solid")


def write_hours(ttl_hours, column_first_week, rmtp_file):
    rmtp_book = openpyxl.load_workbook(filename=rmtp_file)
    rmtp_file_new = fr'{os.path.dirname(rmtp_file)}/upd_{os.path.basename(rmtp_file).split(".")[0]}.xlsx        '
    rmtp_sh = rmtp_book['Evikit']
    rmtp_list_wrk = []
    week_hours = list()

    for row in rmtp_sh.rows:
        rmtp_name = row[0]

        if rmtp_name.value is None:
            pass
        else:

            for i in range(5): #количество столбцов с неделями
                r = re.compile('^[a-zA-Z ]*$')  # паттерн для имени, строка содержит только буквы и пробелы
                if len(rmtp_name.value.split()) >= 2 and r.match(rmtp_name.value):
                    if rmtp_name.value.strip() in ttl_hours.keys():
                        fs_time = ttl_hours.get(rmtp_name.value.strip())
                        cell = rmtp_sh.cell(row=rmtp_name.row, column=column_first_week + i * 2)
                        comm = cell.comment
                        processing_comments(cell, comm, fs_time[i], rmtp_name)
                        cell.value = fs_time[i]
                        cell.number_format = '0.0'
                        rmtp_list_wrk.append(rmtp_name.value.strip())  # добавляем в список обработанного сотрудника
                        week_hours.append(cell.value)
    print('Занесено часов:', sum(week_hours))
    print('Всего занесено в таблицу сотрудников:', len(set(rmtp_list_wrk)))
    for k, v in ttl_hours.items():
        if k not in set(rmtp_list_wrk):
            print(f'Сотрудник есть в часах, но нет в зарплате:', k, v)
    rmtp_book.save(rmtp_file)


def main():
    fs_file, stma_file, rmtp_file = define_of_files(your_target_folder)
    just_open(fs_file)
    just_open(stma_file)
    fs_hours = read_hours(fs_file)
    #pprint.pprint(fs_hours)
    st_hours = read_hours(stma_file)
    #pprint.pprint(st_hours)
    print('***часы FS***')
    write_hours(fs_hours, 5, rmtp_file)

    print('***часы STEMA***')
    write_hours(st_hours, 4, rmtp_file)


if __name__ == "__main__":
    main()

