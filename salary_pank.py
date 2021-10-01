from openpyxl import load_workbook
from yattag import Doc, indent
import datetime
import pandas as pd

your_target_folder = rf'C:\Users\docha\OneDrive\Leka'
my_file = rf'{your_target_folder}\Union Salary\salary_template.xlsx'
company_name = 'UnionCompany OU'
our_iban = ''
our_bic = ''

#wb = load_workbook(my_file)
#ws = wb.worksheets[0]

df = pd.read_excel(my_file)
print(df.columns)
print(sum(df['Total']), df['Total'].count())

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

xml_header = "<?xml version='1.0' encoding='UTF-8'?>"

xml_schema = "<Document xmlns='urn:iso:std:iso:20022:tech:xsd:pain.001.001.03' xmlns:xsi='http://www.w3.org/2001/XMLSchema-instance'>"

xml_cst = '<CstmrCdtTrfInitn>'

xml_grph_o = '<GrpHdr>'

#d = input("Введите сегодняшнюю дату в формате dd.mm.yy: ")
d_now = datetime.datetime.now()
d_00 = d_now.strftime('%d.%m.%y')
xml_date = rf'<MsgId>{d_00}</MsgId>'

d_01 = d_now.strftime('%Y-%m-%d')
xml_date_01 = rf'<CreDtTm>{d_01}T00:00:00.00</CreDtTm>'

nboftxs = df['Total'].count()
xml_nboftxs = f'<NbOfTxs>{nboftxs}</NbOfTxs>'

ctrl_sum = sum(df['Total'])
xml_ctrl = f'<CtrlSum>{ctrl_sum:.2f}</CtrlSum>'

xml_init_o = '<InitgPty>'

xml_company = f'<Nm>{company_name}</Nm>'

xml_init_c = '</InitgPty>'

xml_grph_c = '</GrpHdr>'

doc.asis(xml_header, xml_schema, xml_cst,
         xml_grph_o, xml_date, xml_date_01,
         xml_nboftxs, xml_ctrl,
         xml_init_o, xml_company, xml_init_c,
         xml_grph_c)

xml_header_maksekorraldus = (f'''<PmtInf>
<PmtInfId>PMTID001</PmtInfId>
<PmtMtd>TRF</PmtMtd>
<NbOfTxs>1</NbOfTxs>
<CtrlSum>{}</CtrlSum>
<PmtTpInf>
<SvcLvl>
<Cd>SEPA</Cd>
</SvcLvl>
</PmtTpInf>''')
doc.asis(xml_header_maksekorraldus)

# with tag('<PmtInf>'):
#     for row in ws.iter_rows(min_row=2, max_row=12, min_col=1, max_col=11):
#         row = [cell.value for cell in row]
#         print(row[0], row)
#         with tag('CtrlSum'):
#             row[0]



# with tag('Babies'):
#     # Use ws.max_row for all rows
#     for row in ws.iter_rows(min_row=2, max_row=100, min_col=1, max_col=5):
#         row = [cell.value for cell in row]
#         with tag("Baby"):
#             with tag("Name"):
#                 text(row[2])
#             with tag("Gender"):
#                 text(row[1])
#             with tag("year"):
#                 text(row[0])
#             with tag("count"):
#                 text(row[3])
#             with tag("rank"):
#                 text(row[4])
#
# !СООБЩИТЬ "<?xml version='1.0' encoding='UTF-8'?>"
# !СООБЩИТЬ "<Document xmlns='urn:iso:std:iso:20022:tech:xsd:pain.001.001.03' xmlns:xsi='http://www.w3.org/2001/XMLSchema-instance'>"
# !СООБЩИТЬ "<CstmrCdtTrfInitn>"
# !СООБЩИТЬ "<GrpHdr>"
# !СООБЩИТЬ  "<MsgId>"+РДАТА+"</MsgId>"
# !П_годС=""+"20"+выд(ЗН2\7\2)
# !П_месС=""+выд(ЗН2\4\2)
# !П_денС=""+выд(ЗН2\1\2)
# !П_Date=""+П_годС+"-"+П_месС+"-"+П_денС
# !СООБЩИТЬ "<CreDtTm>"+П_Date+"T00:00:00.00</CreDtTm>"
# !Е0
# !СООБЩИТЬ "<NbOfTxs>"+ИТОГ9+"</NbOfTxs>"
# !Е2
# !СООБЩИТЬ "<CtrlSum>"+ИТОГ7+"</CtrlSum>"
# !СООБЩИТЬ "<InitgPty>"
# !СООБЩИТЬ "<Nm>"+К1+" "+К2+"</Nm>"
# !СООБЩИТЬ "</InitgPty>"
# !СООБЩИТЬ "</GrpHdr>"
# !ВЫБРАТЬ СТРОКИ
# !СООБЩИТЬ "<PmtInf>"
# !СООБЩИТЬ "<PmtInfId>PMTID001</PmtInfId>"
# !СООБЩИТЬ "<PmtMtd>TRF</PmtMtd>"
# !Е0
# !СООБЩИТЬ "<NbOfTxs>"+ИТОГ9+"</NbOfTxs>"
# !Е2
# !СООБЩИТЬ "<CtrlSum>"+ИТОГ7+"</CtrlSum>"
# !СООБЩИТЬ "<PmtTpInf>"
# !СООБЩИТЬ "<SvcLvl>"
# !СООБЩИТЬ "<Cd>SEPA</Cd>"
# !СООБЩИТЬ "</SvcLvl>"
# !СООБЩИТЬ "</PmtTpInf>"
# !П_годС=""+"20"+выд(ЗН{|8}\7\2)
# !П_месС=""+выд(ЗН{|8}\4\2)
# !П_денС=""+выд(ЗН{|8}\1\2)
# !П_Date=""+П_годС+"-"+П_месС+"-"+П_денС
# !СООБЩИТЬ "<ReqdExctnDt>"+П_Date+"</ReqdExctnDt>"
# !СООБЩИТЬ "<Dbtr>"
# !СООБЩИТЬ "<Nm>"+К1+" "+К2+"</Nm>"
# !СООБЩИТЬ "</Dbtr>"
# !СООБЩИТЬ "<DbtrAcct>"
# !СООБЩИТЬ "<Id>"
# !СООБЩИТЬ "<IBAN>"+К21+"</IBAN>"
# !СООБЩИТЬ "</Id>"
# !СООБЩИТЬ "<Ccy>EUR</Ccy>"
# !СООБЩИТЬ "</DbtrAcct>"
# !СООБЩИТЬ "<DbtrAgt>"
# !СООБЩИТЬ "<FinInstnId>"
# !СООБЩИТЬ "<BIC>"+К22+"</BIC>"
# !СООБЩИТЬ "</FinInstnId>"
# !СООБЩИТЬ "</DbtrAgt>"
# !СООБЩИТЬ "<CdtTrfTxInf>"
# !СООБЩИТЬ "<PmtId>"
# !СООБЩИТЬ "<InstrId>"+ЗН{|1}+"</InstrId>"
# !СООБЩИТЬ "<EndToEndId>"+ЗН{|1}+"</EndToEndId>"
# !СООБЩИТЬ  "</PmtId>"
# !СООБЩИТЬ "<Amt>"
# !СООБЩИТЬ "<InstdAmt Ccy='EUR'>"+ЗН{|7}+"</InstdAmt>"
# !СООБЩИТЬ "</Amt>"
# !СООБЩИТЬ "<Cdtr>"
# !СООБЩИТЬ "<Nm>"+ЗН{|4}+"</Nm>"
# !СООБЩИТЬ "</Cdtr>"
# !СООБЩИТЬ "<CdtrAcct>"
# !СООБЩИТЬ "<Id>"
# !ЕСЛИ ЗН|10=1
# !СООБЩИТЬ "<IBAN>"+К401+"</IBAN>"
# !КОНЕЦЕСЛИ
# !ЕСЛИ (ЗН|10=0) И (СЧСВ{|2} <> "2")
# !*СООБЩИТЬ "<IBAN>"+СН{|3}.3^1+"</IBAN>"
# !СООБЩИТЬ "<IBAN>"+СН{|3}.31+"</IBAN>"
# !КОНЕЦЕСЛИ
# !ЕСЛИ (ЗН|10=0) И (СЧСВ{|2} = "2")
# !СООБЩИТЬ "<IBAN>"+СН{|3}.31+"</IBAN>"
# !КОНЕЦЕСЛИ
# !СООБЩИТЬ "</Id>"
# !СООБЩИТЬ "</CdtrAcct>"
# !СООБЩИТЬ "<RmtInf>"
# !СООБЩИТЬ "<Ustrd>"+ЗН{|5}+"</Ustrd>"
# !ЕСЛИ ЗН|11>0
# !СООБЩИТЬ "<Strd>"
# !СООБЩИТЬ "<CdtrRefInf>"
# !СООБЩИТЬ "<Ref>"+ЗН|11+"</Ref>"
# !СООБЩИТЬ "</CdtrRefInf>"
# !СООБЩИТЬ "</Strd>"
# !КОНЕЦЕСЛИ
# !СООБЩИТЬ "</RmtInf>"
# !СООБЩИТЬ "</CdtTrfTxInf>"
# !СООБЩИТЬ "</PmtInf>"
# !СЛЕДУЮЩАЯ СТРОКА
# !СООБЩИТЬ "</CstmrCdtTrfInitn>"
# !СООБЩИТЬ "</Document>"
#
#
#
result = indent(
    doc.getvalue(),
    indentation='   ',
    indent_text=True
)

with open(fr'{your_target_folder}/Union Salary/pank.xml', "w") as f:
    f.write(result)
