#pip install fillpdf
#and poppler dependency conda install -c conda-forge poppler


from fillpdf import fillpdfs
import csv
from collections import OrderedDict
import openpyxl
import os
import config_pdf

#в модуле fillpdfs.py изменить строку 32 с
#data_dict = {}
#на
#data_dict = OrderedDict()
#чтобы получить словарь не в алфавитном порядке, а по мере появления полей

print('docha (d), tati (нажать любую клавишу):')
who = input()
if who == 'd':
    your_target_folder = config_pdf.your_target_folder_docha
else:
    your_target_folder = config_pdf.your_target_folder


def create_new_database_csv(b): #функция для создания новой таблицы, используется только один раз
    #ключи - это первая строка (заголовки), значения - вторая строка
    with open(new_csv, 'w', encoding="utf-8-sig", newline='') as output:
        #encoding sig чтобы при открытии ТОЖЕ отражались все эст.буквы экселем
        w = csv.DictWriter(output, b.keys())
        w.writeheader()
        w.writerow(b)


def fill_pdf_form(wb_, sheet_, origin_pdf, new_pdf, file_name, all_records, row_start, row_last):
    wb = openpyxl.load_workbook(wb_, data_only=True)
    sheet = wb[sheet_]
    data_dict_ = OrderedDict()

    if all_records == 1:
        row_start = 1
        row_last = sheet.max_row
    else:
        row_start = row_start
        row_last = row_last + 1
    for row_nr in range(row_start + 1, row_last + 1):
        date_ = sheet.cell(row=row_nr, column=4).value
        a = 0
        if date_ and not isinstance(date_, int):
            print(date_, type(date_))
            a = len(str(date_))
        if sheet.cell(row=row_nr, column=1).value != '#VALUE!' and a > 7:
            for col_nr in range(1, sheet.max_column + 1):
                head = sheet.cell(row=1, column=col_nr).value
                value = sheet.cell(row=row_nr, column=col_nr).value
                #print(value, type(value))
                if value is None and head not in ['Sugu 3', 'Suhe 3', 'Sugu 2', 'Suhe 2']:
                    value = ''
                data_dict_[head] = value
                # print(head, value)
            # pprint.pprint(data_dict)

            #date_name = date_.strftime('%Y-%m')
            date_name = date_.split('.')[2][2:] + '-' + date_.split('.')[1]
            name = sheet.cell(row=row_nr, column=2).value
            family_name = sheet.cell(row=row_nr, column=3).value

            if not os.path.isdir(f'{new_pdf}/{family_name}_{name}'):
                try:
                    os.mkdir(f'{new_pdf}/{family_name}_{name}')
                except:
                    pass
            fillpdfs.write_fillable_pdf(origin_pdf, f'{new_pdf}/{family_name}_{name}/{file_name}_{date_name}_{name}_{family_name}.pdf', data_dict_)
            # If you want it flattened:
            # fillpdfs.flatten_pdf('new.pdf', 'newflat.pdf')
            print(f'created {new_pdf}/{family_name}_{name}/{file_name}_{date_name}_{name}_{family_name}.pdf')


#*** часть для создания новой таблицы с данными. используется только один раз
#b = fillpdfs.get_form_fields(r'C:\Users\docha\OneDrive\Leka\PPA Elamisluba\Andrii Rudenko\ESTandmedperekonnaliikmetekohta_EST Rudenko.pdf')
#new_csv = r'C:\Users\docha\OneDrive\Leka\PPA Elamisluba\TEtaotlus_EST.csv'
#pprint.pprint(b)
#print(len(b))
#create_new_database_csv(b)
#***конец части для создания новой таблицы


print('выбрать все записи из таблицы? да - 1, нет - 0: ')
all_records = int(input())
if all_records == 1:
    row_start = 1
    row_last = 1
else:
    print('введите номер первой записи: ')
    row_start = int(input())
    print('введите номер последней записи: ')
    row_last = int(input())

excel_book = f'{your_target_folder}\database.xlsx'
new_pdf = f'{your_target_folder}\pdf'

excel_sheet = 'kutse'
origin_pdf = f'{your_target_folder}\Tццandjakutse_EST_20190221.pdf'
file_name = 'Tццandjakutse'
fill_pdf_form(excel_book, excel_sheet, origin_pdf, new_pdf, file_name, all_records, row_start, row_last)


excel_sheet = 'taotlus'
origin_pdf = f'{your_target_folder}\TEtaotlus_EST_20181108.pdf'
file_name = 'Tahtajaliseelamisloataotlus'
fill_pdf_form(excel_book, excel_sheet, origin_pdf, new_pdf, file_name, all_records, row_start, row_last)


excel_sheet = 'pere'
origin_pdf = f'{your_target_folder}\ESTandmedperekonnaliikmetekohta_EST.pdf'
file_name = 'Andmedperekonnaliikmetekohta'
fill_pdf_form(excel_book, excel_sheet, origin_pdf, new_pdf, file_name, all_records, row_start, row_last)
