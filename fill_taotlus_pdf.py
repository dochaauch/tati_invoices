#pip install fillpdf
#and poppler dependency conda install -c conda-forge poppler

from fillpdf import fillpdfs
import pprint
import csv
from collections import defaultdict, OrderedDict
import openpyxl

#в модуле fillpdfs.py изменить строку 32 с
#data_dict = {}
#на
#data_dict = OrderedDict()
#чтобы получить словарь не в алфавитном порядке, а по мере появления полей


def create_new_database_csv(b): #функция для создания новой таблицы, используется только один раз
    #ключи - это первая строка (заголовки), значения - вторая строка
    with open(new_csv, 'w', encoding="utf-8-sig", newline='') as output:
        #encoding sig чтобы при открытии ТОЖЕ отражались все эст.буквы экселем
        w = csv.DictWriter(output, b.keys())
        w.writeheader()
        w.writerow(b)


example_pdf = r'C:\Users\docha\OneDrive\Leka\PPA Elamisluba\Tццandjakutse_EST_20190221.pdf'
new_pdf = r'C:\Users\docha\OneDrive\Leka\PPA Elamisluba\new'

a = fillpdfs.get_form_fields(example_pdf)
#pprint.pprint(a)
#print(len(a))

#*** часть для создания новой таблицы с данными. используется только один раз
#b = fillpdfs.get_form_fields(r'C:\Users\docha\OneDrive\Leka\PPA Elamisluba\ESTtahtajaliseelamisloataotlus_EST Rudenko.pdf')
#new_csv = r'C:\Users\docha\OneDrive\Leka\PPA Elamisluba\TEtaotlus_EST.csv'
#pprint.pprint(b)
#print(len(b))
#create_new_database_csv(b)
#***конец части для создания новой таблицы


# returns a dictionary of fields
# Set the returned dictionary values a save to a variable
# For radio boxes ('Off' = not filled, 'Yes' = filled)


wb = openpyxl.load_workbook(r'C:\Users\docha\OneDrive\Leka\PPA Elamisluba\database.xlsx', data_only=True)
sheet = wb['Tooandjakutse']
data_dict = OrderedDict()
for row_nr in range(2, sheet.max_row + 1):
    for col_nr in range(1, sheet.max_column + 1):
        head = sheet.cell(row = 1, column = col_nr).value
        value = sheet.cell(row = row_nr, column = col_nr).value
        if value is None:
            value = ''
        data_dict[head] = value
        #print(head, value)
    #pprint.pprint(data_dict)
    name = sheet.cell(row=row_nr, column=3).value
    fillpdfs.write_fillable_pdf(example_pdf, f'{new_pdf}_{name}.pdf', data_dict)
    print(f'created {new_pdf}_{name}.pdf')



data_dict = {
'Eesnimi või -nimed': 'Anatoli',
'Isikukood või sünniaeg': '47901050281',
'alus': 0,
}

#fillpdfs.write_fillable_pdf(example_pdf, new_pdf, data_dict)

# If you want it flattened:
#fillpdfs.flatten_pdf('new.pdf', 'newflat.pdf')