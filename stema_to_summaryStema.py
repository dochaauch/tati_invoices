import openpyxl
import xlrd
import os
import pprint
import re
from openpyxl.comments import Comment
import hours_config

print('docha (d), tati (нажать любую клавишу):')
who = input()
if who == 'd':
    your_target_folder = hours_config.your_target_folder_docha
else:
    your_target_folder = hours_config.your_target_folder



def define_of_files(your_target_folder):
    for dirpath, _, filenames in os.walk(your_target_folder):
        fs_file_ = re.compile(r'.*(FS).*')
        stma_file_ = re.compile(r'.*(STEMA).*')
        rmtp_file_ = re.compile(r'^(rmtp).*')
        for items in filenames:
            if items.lower().endswith('.xlsx'):
                fs_file__ = fs_file_.match(items)
                if fs_file__:
                    fs_file = fr'{your_target_folder}/{fs_file__[0]}'

                stma_file__ = stma_file_.match(items)
                if stma_file__:
                    stma_file = fr'{your_target_folder}/{stma_file__[0]}'

                rmtp_file__ = rmtp_file_.match(items)
                if rmtp_file__:
                    rmtp_file = fr'{your_target_folder}/{rmtp_file__[0]}'
    return fs_file, stma_file, rmtp_file





def list_of_files(your_target_folder):
    week_files = []
    week_list = []
    for dirpath, _, filenames in os.walk(your_target_folder):
        r = re.compile(r'.*Vko(\d{2}).*')
        for items in filenames:
            if items.lower().startswith('finstratum vko'):
                file_full_path = os.path.abspath(os.path.join(dirpath, items))
                week_nr = r.match(items)[1]
                week_files.append(file_full_path)
                week_list.append(week_nr)
    return week_files, week_list


def read_week_file(week_file):
    week_book = xlrd.open_workbook(filename=week_file)
    week_hours = dict()
    sh = week_book.sheet_by_index(0)
    head_row = 1
    first_day = 1
    last_day = 7
    print(os.path.basename(week_file))
    pprint.pprint(sh.row_values(head_row)[first_day - 1: last_day + 2])
    for rx in range(head_row + 1, sh.nrows):
        fs_name = sh.row(rx)[0]
        if len(fs_name.value.split()) >= 2:
            range_hours = sh.row(rx)[first_day: last_day + 1]
            if fs_name.value:
                list_hours = []
                for l_ in range_hours:
                    list_hours.append(l_.value)
                week_hours[fs_name.value.strip()] = list_hours
    return week_hours


def processing_comments(cell, comm, stema_time_i, stema_name):
    if cell.value:
        if cell.value != stema_time_i and not (cell.value is None or cell.value == 0):
            if comm:  # если у ячейки уже был комментарий
                comm = Comment(f'{comm.text}, было: {cell.value}, стало: {stema_time_i}', 'TJ')
            else:  # если комментов не было, но предыдущее значение ячейки не 0 и не None
                comm = Comment(f'было: {cell.value}, стало: {stema_time_i}', 'TJ')
                print(stema_name.value, cell.coordinate, comm.text)
                cell.comment = comm
                cell.fill = openpyxl.styles.PatternFill(start_color='FFC7CE', fill_type="solid")


def write_hours_to_summary(week_hours, column_day, stema_file):
    stema_book = openpyxl.load_workbook(filename=stema_file)
    stema_sheets = stema_book.sheetnames[0]
    stema_sh = stema_book[stema_sheets]
    stema_list_wrk = []
    words = ['Total h', 'Extra h']
    chk_pat = '(?:{})'.format('|'.join(words))
    for row in stema_sh.rows:
        stema_name = row[0]
        if row[0].value is None:
            pass
        else:
            r = re.compile('^[a-zA-Z ]*$')  # паттерн для имени, строка содержит только буквы и пробелы
            if (len(stema_name.value.split()) >= 2 and r.match(stema_name.value)
                    and not bool(re.search(chk_pat, stema_name.value, flags=re.I))):
                if stema_name.value.strip() in week_hours.keys():
                    stema_time = week_hours.get(stema_name.value.strip())
                    for i in range(7):  # количество колонок с днями недели
                        cell = stema_sh.cell(row=stema_name.row, column=column_day + i)  # перебираем дни недели
                        comm = cell.comment
                        processing_comments(cell, comm, stema_time[i], stema_name)  # обрабатываем комментарии
                        cell.value = stema_time[i]
                        stema_list_wrk.append(stema_name.value.strip())  # добавляем в список обработанного сотрудника
    print('Всего занесено в таблицу сотрудников:', len(set(stema_list_wrk)))
    for k, v in week_hours.items():
        if k not in set(stema_list_wrk) and not bool(re.search(chk_pat, k, flags=re.I)):
            print(f'Сотрудник есть в неделях, но нет в сводной:', k, v)
    stema_book.save(stema_file)


def main():
    fs_file, stema_file, rmtp_file = define_of_files(your_target_folder)
    week_files, week_list = list_of_files(your_target_folder)
    print(week_list)
    for cnt_f, w_file in enumerate(week_files):
        week_hours = read_week_file(w_file)
        write_hours_to_summary(week_hours, 2 + cnt_f * 7, stema_file)


if __name__ == "__main__":
    main()
