import openpyxl
import xlrd
import os
import pprint
import re
from openpyxl.comments import Comment
import hours_config

print('docha (d), tati (нажать любую клавишу):')
who = input()
if who == 'd':
    your_target_folder = hours_config.your_target_folder_docha
else:
    your_target_folder = hours_config.your_target_folder



def define_of_files(your_target_folder):
    for dirpath, _, filenames in os.walk(your_target_folder):
        fs_file_ = re.compile(r'.*(FS).*')
        stma_file_ = re.compile(r'.*(STEMA).*')
        rmtp_file_ = re.compile(r'^(rmtp).*')
        for items in filenames:
            if items.lower().endswith('.xlsx'):
                fs_file__ = fs_file_.match(items)
                if fs_file__:
                    fs_file = fr'{your_target_folder}/{fs_file__[0]}'

                stma_file__ = stma_file_.match(items)
                if stma_file__:
                    stma_file = fr'{your_target_folder}/{stma_file__[0]}'

                rmtp_file__ = rmtp_file_.match(items)
                if rmtp_file__:
                    rmtp_file = fr'{your_target_folder}/{rmtp_file__[0]}'
    return fs_file, stma_file, rmtp_file





def list_of_files(your_target_folder):
    week_files = []
    week_list = []
    for dirpath, _, filenames in os.walk(your_target_folder):
        r = re.compile(r'.*Vko(\d{2}).*')
        for items in filenames:
            if items.lower().startswith('finstratum vko'):
                file_full_path = os.path.abspath(os.path.join(dirpath, items))
                week_nr = r.match(items)[1]
                week_files.append(file_full_path)
                week_list.append(week_nr)
    return week_files, week_list


def change_time_to_norm(str_):
    if str_.value:
        return float(str(str_.value).replace('.3', '.5'))


def calculate_hours(start_hour, end_hour, week_nr, day_name, fs_name):
    if end_hour > start_hour:
        work_hour = end_hour - start_hour - 0.5
    else:
        work_hour = 24 - start_hour + end_hour - 0.5
        print('*** ПОЛУНОЧНИК ***'.center(36), week_nr, day_name, fs_name.value.ljust(30),
              start_hour, end_hour, work_hour)
    return work_hour


def read_every_page_week_file(week_file):
    list_of_week_hours = []
    week_book = xlrd.open_workbook(filename=week_file)
    r = re.compile(r'.*Vko(\d{2}).*')  # выделяем номер недели в названии файла
    week_nr = r.match(week_file)[1]
    for i in range(1, 8):  # перебираем по дню листы каждой книги недели
        sh = week_book.sheet_by_index(i)
        head_row = 1
        week_name_dict = {1: 'Пн', 2: 'Вт', 3: 'Ср', 4: 'Чт', 5: 'Пт', 6: 'Сб', 7: 'Вс'}
        day_name = week_name_dict.get(i)
        for rx in range(head_row + 1, sh.nrows):
            rate = sh.row(rx)[5]
            remarks = sh.row(rx)[10]
            r = re.compile(r'.*(\b(\d+)h\s*FS).*')  # ищем в примечании часы FS
            fs_name = sh.row(rx)[0]
            start_hour = change_time_to_norm(sh.row(rx)[3])
            end_hour = change_time_to_norm(sh.row(rx)[4])
            remark = ''
            add_hours = 0
            less_hours = 0
            if rate.value != 'FS' and rate.value and start_hour and end_hour:
                work_hour = calculate_hours(start_hour, end_hour, week_nr, day_name, fs_name)
                less_hours = work_hour - rate.value
                if work_hour != rate.value:
                    print('+++ часы не сходятся с расчетом +++'.center(36), week_nr, day_name, fs_name.value.ljust(30),
                          start_hour, end_hour, 'в табл:', rate.value, 'расч:', work_hour,
                          'разн', work_hour - rate.value)

            if rate.value == 'FS' or r.match(remarks.value) or 'foreman' in remarks.value:
                if rate.value == 'FS':
                    work_hour = calculate_hours(start_hour, end_hour, week_nr, day_name, fs_name)
                if r.match(remarks.value):
                    add_hours = work_hour = float(r.match(remarks.value).group(2))
                    remark = remarks.value
                    print('... добавлены часы из примечания ...'.center(36), week_nr, day_name, fs_name.value.ljust(30),
                          start_hour, end_hour, 'осн.часы:', rate.value,
                          'доб.часы FS:', work_hour, remarks.value)
                if 'foreman' in remarks.value:
                    if rate.value != 'FS':
                        work_hour = calculate_hours(start_hour, end_hour, week_nr, day_name, fs_name)
                        if work_hour > rate.value:
                            add_hours = work_hour = work_hour - rate.value
                        else:
                            work_hour = 0
                        print('--- работал за мастера ---'.center(36), week_nr, day_name, fs_name.value.ljust(30),
                              start_hour, end_hour, 'осн.часы:', rate.value,
                              'доб.часы FS:', work_hour, remarks.value)

                dict_of_hours = {'name': fs_name.value, 'week_nr': week_nr, 'day_name': day_name,
                                 'work_hour': work_hour, 'remark': remark}
                list_of_week_hours.append(dict_of_hours)
            if not less_hours - add_hours == 0:
                print('<<< уточнить, часы не сходятся! >>>'.center(36), week_nr, day_name, fs_name.value.ljust(30),
                      start_hour, end_hour, 'разница', less_hours - add_hours)
    return list_of_week_hours


def processing_comments(cell, comm, stema_time_i, stema_name):
    if comm:
        if cell.value != stema_time_i:

            if comm:  # если у ячейки уже был комментарий
                comm = Comment(f'{comm}, было: {cell.value}, стало: {stema_time_i}', 'TJ')
            else:  # если комментов не было, но предыдущее значение ячейки не 0 и не None
                comm = Comment(f'было: {comm}, стало: {stema_time_i}', 'TJ')
                print(stema_name.value, cell.coordinate, comm.text)
                cell.comment = comm
                cell.fill = openpyxl.styles.PatternFill(start_color='FFC7CE', fill_type="solid")


def write_hours_to_summary(list_of_week_hours, column_day, summary_file):
    summary_book = openpyxl.load_workbook(filename=summary_file)
    summary_sheets = summary_book.sheetnames[0]
    summary_sh = summary_book[summary_sheets]
    summary_list_wrk = []
    #nr_day = 0
    #words = ['Total h', 'Extra h']
    #chk_pat = '(?:{})'.format('|'.join(words))
    for work_hours in list_of_week_hours:
        for row in summary_sh.rows:
            summary_name = row[0]
            if summary_name.value:
                if summary_name.value.strip() in work_hours['name']:
                    summary_time = work_hours['work_hour']
                    day_name = work_hours['day_name']
                    week_nr_dict = {'Пн': 0, 'Вт': 1, 'Ср': 2, 'Чт': 3, 'Пт': 4, 'Сб': 5, 'Вс': 6}
                    nr_day = week_nr_dict.get(day_name)
                    cell = summary_sh.cell(row=summary_name.row, column=column_day + nr_day)
                    comm = cell.comment
                    if comm is None:
                        comm = work_hours['remark']
                    else:
                        comm = comm.text + work_hours['remark']
                    processing_comments(cell, comm, summary_time, summary_name)  # обрабатываем комментарии
                    cell.value = summary_time
                    summary_list_wrk.append(summary_name.value.strip())  # добавляем в список обработанного сотрудника
                #else:
                #    print('Сотрудника нет в сводной!', work_hours['name'])

    print('Всего занесено в таблицу сотрудников:', len(set(summary_list_wrk)))
    summ_l = list()
    for row in summary_sh.rows:
        if row[0].value:
            summ_l.append(row[0].value.strip())
    for slw in summary_list_wrk:
        if slw not in set(summ_l):
            print(f'Сотрудник есть в неделях, но нет в сводной:', slw)
    summary_book.save(summary_file)


def main():
    summary_file, stma_file, rmtp_file = define_of_files(your_target_folder)
    week_files, week_list = list_of_files(your_target_folder)
    for cnt_f, w_file in enumerate(week_files):
        list_of_week_hours = read_every_page_week_file(w_file)
        write_hours_to_summary(list_of_week_hours, 2 + cnt_f * 7, summary_file)



if __name__ == "__main__":
    main()

